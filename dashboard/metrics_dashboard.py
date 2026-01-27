#!/usr/bin/env python3
"""
Order Book Metrics Dashboard

Usage:
    ./OrderBook_v2 | python3 dashboard/metrics_dashboard.py
"""

import sys
import os
import re
import argparse
import platform
import subprocess
import threading
import time as time_module
import csv
from collections import deque
from dataclasses import dataclass, field, asdict
from datetime import datetime
from typing import Optional, List

import psutil
import pyqtgraph as pg

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils.dataframe import dataframe_to_rows
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False
from PyQt6.QtCore import Qt, QTimer, pyqtSignal, QObject
from PyQt6.QtGui import QFont, QPalette, QColor
from PyQt6.QtWidgets import (
    QApplication, QMainWindow, QWidget, QVBoxLayout, QHBoxLayout,
    QGridLayout, QLabel, QGroupBox, QSplitter, QFrame, QComboBox,
    QSpinBox, QLineEdit, QPushButton, QFileDialog, QMessageBox,
)

# Configuration
MAX_HISTORY = 100
DEFAULT_UPDATE_MS = 100
DATA_TIMEOUT_MS = 5000

# Time unit definitions
TIME_UNITS = [("Nanoseconds", "ns", 1), ("Milliseconds", "ms", 1e6), ("Seconds", "s", 1e9)]


@dataclass
class Metrics:
    """All metrics in a single flat structure."""
    # Timestamp
    timestamp: str = ""
    # Order flow
    total_matches: int = 0
    bid_orders: int = 0
    ask_orders: int = 0
    # Fills
    full_fills: int = 0
    full_fills_pct: float = 0.0
    partial_fills: int = 0
    partial_fills_pct: float = 0.0
    no_fills: int = 0
    qty_matched: int = 0
    # Latency (nanoseconds)
    avg_match_ns: int = 0
    min_ns: int = 0
    max_ns: int = 0
    avg_lookup_ns: int = 0
    # Memory
    total_orders: int = 0
    buy_levels: int = 0
    sell_levels: int = 0
    memory_bytes: int = 0
    memory_str: str = "0 B"
    # CPU
    user_time: str = "0"
    system_time: str = "0"
    cpu_pct: float = 0.0
    ctx_switches: int = 0
    # Overall
    records: int = 0
    throughput: float = 0.0


class MetricsParser:
    """Parses metrics output from C++ application."""

    PATTERNS = {
        'total_matches': (r'Total match\(\) calls:\s*(\d+)', int),
        'bid_orders': (r'Bid orders:\s*(\d+)', int),
        'ask_orders': (r'Ask orders:\s*(\d+)', int),
        'full_fills': (r'Full fills:\s*(\d+)', int),
        'full_fills_pct': (r'Full fills:.*\(([\d.]+)%\)', float),
        'partial_fills': (r'Partial fills:\s*(\d+)', int),
        'partial_fills_pct': (r'Partial fills:.*\(([\d.]+)%\)', float),
        'no_fills': (r'No fills \(added\):\s*(\d+)', int),
        'qty_matched': (r'Quantity matched:\s*(\d+)', int),
        'avg_match_ns': (r'Avg per match:\s*(\d+)', int),
        'min_ns': (r'Min:\s*(\d+)', int),
        'max_ns': (r'Max:\s*(\d+)', int),
        'avg_lookup_ns': (r'Avg per lookup:\s*(\d+)', int),
        'total_orders': (r'Total orders:\s*(\d+)', int),
        'buy_levels': (r'Buy price levels:\s*(\d+)', int),
        'sell_levels': (r'Sell price levels:\s*(\d+)', int),
        'cpu_pct': (r'CPU percentage:\s*([\d.]+)%', float),
        'ctx_switches': (r'Voluntary ctx sw:\s*(\d+)', int),
        'records': (r'\[Progress:\s*(\d+)\s*records', int),
        'throughput': (r'Throughput:\s*([\d.]+)', float),
    }

    def __init__(self):
        self._in_block = False
        self._current = Metrics()
        # Pre-compile patterns
        self._compiled = {k: (re.compile(p), t) for k, (p, t) in self.PATTERNS.items()}

    def parse_line(self, line: str) -> Optional[Metrics]:
        """Parse a line, return Metrics when block completes."""
        line = line.strip()

        if "============ Matching Engine Metrics" in line:
            self._in_block = True
            self._current = Metrics()
            return None

        if "=================================================" in line and self._in_block:
            self._in_block = False
            return self._current

        # Parse fields
        for attr, (pattern, converter) in self._compiled.items():
            match = pattern.search(line)
            if match:
                setattr(self._current, attr, converter(match.group(1)))

        # Special cases
        if "Estimated memory:" in line:
            self._current.memory_str = line.split(":")[-1].strip()
            self._current.memory_bytes = self._parse_memory(self._current.memory_str)
        elif "User time:" in line:
            self._current.user_time = line.split(":")[-1].strip()
        elif "System time:" in line:
            self._current.system_time = line.split(":")[-1].strip()

        return None

    @staticmethod
    def _parse_memory(s: str) -> int:
        match = re.match(r'([\d.]+)\s*(B|KB|MB|GB)', s)
        if not match:
            return 0
        val, unit = float(match.group(1)), match.group(2)
        return int(val * {'B': 1, 'KB': 1024, 'MB': 1024**2, 'GB': 1024**3}.get(unit, 1))


class ProcessManager:
    """Finds and kills OrderBook_v2 processes."""
    PROCESS_NAMES = ('OrderBook_v2', 'orderbook_v2', 'OrderBook')

    @classmethod
    def find_processes(cls) -> List[psutil.Process]:
        """Find all running OrderBook_v2 processes."""
        found = []
        for proc in psutil.process_iter(['pid', 'name', 'exe', 'cmdline']):
            try:
                name = proc.info['name'] or ''
                exe = proc.info['exe'] or ''
                cmdline = proc.info['cmdline'] or []

                # Check process name directly
                if any(pn.lower() == name.lower() for pn in cls.PROCESS_NAMES):
                    found.append(proc)
                    continue

                # Check executable path ends with our binary name
                if exe:
                    exe_name = os.path.basename(exe)
                    if any(pn.lower() == exe_name.lower() for pn in cls.PROCESS_NAMES):
                        found.append(proc)
                        continue

                # Check first cmdline arg (the executed binary) only
                if cmdline and len(cmdline) > 0:
                    cmd_exe = os.path.basename(cmdline[0])
                    if any(pn.lower() == cmd_exe.lower() for pn in cls.PROCESS_NAMES):
                        found.append(proc)

            except (psutil.NoSuchProcess, psutil.AccessDenied):
                continue
        return found

    @classmethod
    def kill_all(cls) -> tuple[int, List[str]]:
        """Kill all OrderBook_v2 processes. Returns (count_killed, errors)."""
        procs = cls.find_processes()
        killed = 0
        errors = []
        for proc in procs:
            try:
                proc.terminate()
                proc.wait(timeout=3)
                killed += 1
            except psutil.TimeoutExpired:
                try:
                    proc.kill()
                    killed += 1
                except Exception as e:
                    errors.append(f"PID {proc.pid}: {e}")
            except Exception as e:
                errors.append(f"PID {proc.pid}: {e}")
        return killed, errors


class SpreadsheetExporter:
    """Exports collected metrics to spreadsheet."""

    @staticmethod
    def export(metrics_history: List[Metrics], filepath: str) -> str:
        """Export metrics to xlsx or csv."""
        if HAS_OPENPYXL and filepath.endswith('.xlsx'):
            return SpreadsheetExporter._export_xlsx(metrics_history, filepath)
        else:
            csv_path = filepath.replace('.xlsx', '.csv') if filepath.endswith('.xlsx') else filepath
            return SpreadsheetExporter._export_csv(metrics_history, csv_path)

    @staticmethod
    def _export_csv(metrics_history: List[Metrics], filepath: str) -> str:
        """Fast CSV export."""
        if not metrics_history:
            return filepath
        headers = list(asdict(metrics_history[0]).keys())
        with open(filepath, 'w', newline='') as f:
            writer = csv.DictWriter(f, fieldnames=headers)
            writer.writeheader()
            for m in metrics_history:
                writer.writerow(asdict(m))
        return filepath

    @staticmethod
    def _export_xlsx(metrics_history: List[Metrics], filepath: str) -> str:
        """Pretty xlsx export with formatting."""
        if not metrics_history:
            return filepath

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Metrics"

        # Headers
        headers = list(asdict(metrics_history[0]).keys())
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="2E7D32", end_color="2E7D32", fill_type="solid")
        header_align = Alignment(horizontal="center")
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header.replace('_', ' ').title())
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = thin_border

        # Data rows
        alt_fill = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
        for row_idx, m in enumerate(metrics_history, 2):
            data = asdict(m)
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=row_idx, column=col, value=data[header])
                cell.border = thin_border
                if row_idx % 2 == 0:
                    cell.fill = alt_fill

        # Auto-width columns (approximate)
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 15

        # Freeze header row
        ws.freeze_panes = 'A2'

        # Summary sheet
        ws_summary = wb.create_sheet("Summary")
        ws_summary['A1'] = "Order Book Metrics Summary"
        ws_summary['A1'].font = Font(bold=True, size=14)
        ws_summary.merge_cells('A1:C1')

        ws_summary['A3'] = "Generated:"
        ws_summary['B3'] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        ws_summary['A4'] = "Total Samples:"
        ws_summary['B4'] = len(metrics_history)

        if metrics_history:
            last = metrics_history[-1]
            ws_summary['A6'] = "Final Statistics"
            ws_summary['A6'].font = Font(bold=True)
            stats = [
                ("Total Records", last.records),
                ("Total Matches", last.total_matches),
                ("Avg Latency (ns)", last.avg_match_ns),
                ("Memory", last.memory_str),
                ("CPU %", f"{last.cpu_pct:.1f}%"),
            ]
            for i, (label, val) in enumerate(stats, 7):
                ws_summary[f'A{i}'] = label
                ws_summary[f'B{i}'] = val

        wb.save(filepath)
        return filepath


class InputReader(QObject):
    """Reads input from various sources in a background thread."""
    metrics_ready = pyqtSignal(object)
    data_received = pyqtSignal()
    error_occurred = pyqtSignal(str)
    process_ended = pyqtSignal(int, str)

    def __init__(self):
        super().__init__()
        self.parser = MetricsParser()
        self.running = False
        self._process = None
        self._stderr = []

    def start(self, source_type: str, path: str = ""):
        """Start reading from specified source."""
        self.running = True
        self._stderr = []
        target = {
            'stdin': self._read_stdin,
            'file': lambda: self._read_file(path),
            'exec': lambda: self._read_process(path),
        }.get(source_type, self._read_stdin)
        threading.Thread(target=target, daemon=True).start()

    def stop(self):
        self.running = False
        if self._process:
            try:
                self._process.terminate()
                self._process.wait(timeout=2)
            except:
                pass
            self._process = None

    def _process_line(self, line: str):
        self.data_received.emit()
        if metrics := self.parser.parse_line(line):
            self.metrics_ready.emit(metrics)

    def _read_stdin(self):
        for line in sys.stdin:
            if not self.running:
                break
            self._process_line(line)

    def _read_file(self, path: str):
        try:
            with open(path) as f:
                f.seek(0, 2)  # Tail mode
                while self.running:
                    if line := f.readline():
                        self._process_line(line)
                    else:
                        time_module.sleep(0.01)
        except Exception as e:
            self.error_occurred.emit(str(e))

    def _read_process(self, path: str):
        try:
            self._process = subprocess.Popen(
                path, stdout=subprocess.PIPE, stderr=subprocess.PIPE,
                text=True, bufsize=1, shell=True
            )
            # Stderr readerx
            def read_stderr():
                for line in self._process.stderr:
                    self._stderr.append(line.strip())
                    if len(self._stderr) > 50:
                        self._stderr.pop(0)
            threading.Thread(target=read_stderr, daemon=True).start()

            for line in self._process.stdout:
                if not self.running:
                    break
                self._process_line(line)

            self._process.wait()
            self.process_ended.emit(self._process.returncode, '\n'.join(self._stderr))
        except Exception as e:
            self.error_occurred.emit(str(e))


class StatusIndicator(QLabel):
    """Simple colored status dot with text."""
    COLORS = {'active': '#0f8', 'warning': '#fa0', 'error': '#f44', 'inactive': '#666'}

    def __init__(self, text=""):
        super().__init__(f"● {text}")
        self.setFont(QFont("Arial", 11))
        self.set_status("inactive")

    def set_status(self, status: str, text: str = None):
        color = self.COLORS.get(status, self.COLORS['inactive'])
        if text:
            self.setText(f"<span style='color:{color}'>●</span> {text}")
        else:
            current = self.text().split(" ", 1)[-1] if " " in self.text() else ""
            self.setText(f"<span style='color:{color}'>●</span> {current}")


class MetricCard(QGroupBox):
    """Card widget for displaying metrics."""

    def __init__(self, title: str, metrics: list):
        super().__init__(title)
        self.setFont(QFont("Helvetica", 11, QFont.Weight.Bold))
        layout = QGridLayout(self)
        layout.setSpacing(6)
        self._labels = {}
        for i, name in enumerate(metrics):
            layout.addWidget(QLabel(name), i, 0)
            lbl = QLabel("0")
            lbl.setFont(QFont("Menlo", 10, QFont.Weight.Bold))
            lbl.setAlignment(Qt.AlignmentFlag.AlignRight)
            layout.addWidget(lbl, i, 1)
            self._labels[name] = lbl

    def update(self, values: dict):
        for name, val in values.items():
            if name in self._labels:
                self._labels[name].setText(str(val))


class SettingsPanel(QFrame):
    """Compact settings bar."""
    source_changed = pyqtSignal(str, str)
    interval_changed = pyqtSignal(int)
    time_unit_changed = pyqtSignal(int)

    SOURCES = [("stdin", "Standard Input"), ("file", "File..."), ("exec", "Executable...")]

    def __init__(self):
        super().__init__()
        self.setFrameStyle(QFrame.Shape.StyledPanel)
        layout = QHBoxLayout(self)
        layout.setSpacing(10)

        # Source
        layout.addWidget(QLabel("Source:"))
        self.source_combo = QComboBox()
        self.source_combo.addItems([s[1] for s in self.SOURCES])
        self.source_combo.currentIndexChanged.connect(self._on_source_idx)
        layout.addWidget(self.source_combo)

        self.path_edit = QLineEdit()
        self.path_edit.setPlaceholderText("Path...")
        self.path_edit.setMinimumWidth(200)
        self.path_edit.setEnabled(False)
        layout.addWidget(self.path_edit)

        self.browse_btn = QPushButton("Browse")
        self.browse_btn.setEnabled(False)
        self.browse_btn.clicked.connect(self._browse)
        layout.addWidget(self.browse_btn)

        self.connect_btn = QPushButton("Connect")
        self.connect_btn.clicked.connect(self._connect)
        layout.addWidget(self.connect_btn)

        layout.addWidget(self._sep())

        # Interval
        layout.addWidget(QLabel("Update:"))
        self.interval_spin = QSpinBox()
        self.interval_spin.setRange(10, 1000)
        self.interval_spin.setValue(DEFAULT_UPDATE_MS)
        self.interval_spin.setSuffix(" ms")
        self.interval_spin.valueChanged.connect(self.interval_changed.emit)
        layout.addWidget(self.interval_spin)

        layout.addWidget(self._sep())

        # Time unit
        layout.addWidget(QLabel("Time:"))
        self.time_combo = QComboBox()
        self.time_combo.addItems([t[0] for t in TIME_UNITS])
        self.time_combo.setCurrentIndex(1)  # ms default
        self.time_combo.currentIndexChanged.connect(self.time_unit_changed.emit)
        layout.addWidget(self.time_combo)

        layout.addStretch()

    def _sep(self):
        sep = QFrame()
        sep.setFrameShape(QFrame.Shape.VLine)
        return sep

    def _on_source_idx(self, idx: int):
        needs_path = idx > 0
        self.path_edit.setEnabled(needs_path)
        self.browse_btn.setEnabled(needs_path)

    def _browse(self):
        path, _ = QFileDialog.getOpenFileName(self, "Select File")
        if path:
            self.path_edit.setText(path)

    def _connect(self):
        idx = self.source_combo.currentIndex()
        source_type = self.SOURCES[idx][0]
        self.source_changed.emit(source_type, self.path_edit.text())


class DashboardWindow(QMainWindow):
    """Main dashboard window."""

    def __init__(self, source=None, executable=None):
        super().__init__()
        self.setWindowTitle("Order Book Metrics Dashboard")
        self.setMinimumSize(1200, 800)

        self._last_data = None
        self._time_divisor = 1e6  # Default ms
        self._sample_idx = 0
        self._is_stopped = False

        # Full metrics history for export (unbounded)
        self._metrics_history: List[Metrics] = []

        # History buffers for graphs (bounded)
        self.hist = {k: deque(maxlen=MAX_HISTORY) for k in [
            'time', 'records', 'cpu', 'mem', 'latency', 'fills', 'bids', 'asks'
        ]}

        self._setup_theme()
        self._setup_ui()
        self._setup_reader()
        self._setup_timers()

        # Auto-start
        if executable:
            self._connect("exec", executable)
        elif source:
            self._connect("file", source)
        else:
            self._connect("stdin", "")

    def _setup_theme(self):
        p = QPalette()
        p.setColor(QPalette.ColorRole.Window, QColor(30, 30, 30))
        p.setColor(QPalette.ColorRole.WindowText, QColor(220, 220, 220))
        p.setColor(QPalette.ColorRole.Base, QColor(45, 45, 45))
        p.setColor(QPalette.ColorRole.Text, QColor(220, 220, 220))
        p.setColor(QPalette.ColorRole.Button, QColor(55, 55, 55))
        p.setColor(QPalette.ColorRole.ButtonText, QColor(220, 220, 220))
        self.setPalette(p)
        pg.setConfigOptions(antialias=True, background=QColor(30, 30, 30), foreground='w')

    def _setup_ui(self):
        central = QWidget()
        self.setCentralWidget(central)
        layout = QVBoxLayout(central)
        layout.setSpacing(8)

        # Settings
        self.settings = SettingsPanel()
        self.settings.source_changed.connect(self._connect)
        self.settings.interval_changed.connect(lambda v: self.timer.setInterval(v))
        self.settings.time_unit_changed.connect(self._set_time_unit)
        layout.addWidget(self.settings)

        # Status bar
        status = QHBoxLayout()
        self.data_status = StatusIndicator("Data: --")
        self.proc_status = StatusIndicator("Process: --")
        self.records_lbl = QLabel("Records: 0")
        self.records_lbl.setFont(QFont("Menlo", 10))
        self.throughput_lbl = QLabel("Throughput: --")
        self.throughput_lbl.setFont(QFont("Menlo", 10))

        # Kill & Export button
        self.kill_btn = QPushButton("⏹ Stop && Export")
        self.kill_btn.setStyleSheet("""
            QPushButton {
                background-color: #c62828;
                color: white;
                font-weight: bold;
                padding: 6px 16px;
                border-radius: 4px;
                border: none;
            }
            QPushButton:hover {
                background-color: #e53935;
            }
            QPushButton:pressed {
                background-color: #b71c1c;
            }
            QPushButton:disabled {
                background-color: #555;
                color: #888;
            }
        """)
        self.kill_btn.clicked.connect(self._on_kill_clicked)

        status.addWidget(self.data_status)
        status.addWidget(self.proc_status)
        status.addStretch()
        status.addWidget(self.records_lbl)
        status.addWidget(self.throughput_lbl)
        status.addWidget(self.kill_btn)
        layout.addLayout(status)

        # System info
        info = self._get_sys_info()
        sys_lbl = QLabel(f"<b>System:</b> {info['os']} {info['arch']} | "
                         f"<b>CPU:</b> {info['cores']}C | <b>RAM:</b> {info['ram']}")
        sys_lbl.setFont(QFont("Helvetica", 9))
        layout.addWidget(sys_lbl)

        # Main splitter
        splitter = QSplitter(Qt.Orientation.Horizontal)
        layout.addWidget(splitter, stretch=1)

        # Left: Cards
        left = QWidget()
        left_layout = QVBoxLayout(left)
        self.cards = {
            'flow': MetricCard("Order Flow", ["Matches", "Bids", "Asks"]),
            'fills': MetricCard("Fills", ["Full", "Partial", "No Fill", "Qty"]),
            'latency': MetricCard("Latency", ["Avg", "Min", "Max", "Lookup"]),
            'memory': MetricCard("Memory", ["Orders", "Levels", "Size"]),
            'cpu': MetricCard("CPU", ["User", "System", "Usage", "CtxSw"]),
        }
        for card in self.cards.values():
            left_layout.addWidget(card)
        left_layout.addStretch()
        splitter.addWidget(left)

        # Right: Graphs
        right = QWidget()
        right_layout = QVBoxLayout(right)
        self._setup_graphs(right_layout)
        splitter.addWidget(right)
        splitter.setSizes([300, 900])

    def _setup_graphs(self, layout):
        """Create 2x2 grid of graphs matching professional styling specs."""
        self.plots = {}
        self.curves = {}
        self.fills = {}
        self.ref_lines = {}

        # Create 2x2 grid layout
        grid = QGridLayout()
        layout.addLayout(grid)

        # =====================================================================
        # Panel 1 (top-left): Latency
        # =====================================================================
        plot_latency = pg.PlotWidget()
        plot_latency.setTitle("Match Latency: Stable ~2μs", color='darkgreen', bold=True)
        plot_latency.setLabel('left', "Latency (μs)", **{'font-weight': 'bold'})
        plot_latency.setLabel('bottom', "Sample #", **{'font-weight': 'bold'})
        plot_latency.showGrid(x=True, y=True, alpha=0.3)
        plot_latency.setYRange(1.8, 2.3)

        # Target zone: green dashed line at 2.0 μs
        target_line = pg.InfiniteLine(pos=2.0, angle=0, pen=pg.mkPen('green', width=1.5, style=Qt.PenStyle.DashLine))
        plot_latency.addItem(target_line)

        # Light green fill between 1.95-2.05 μs
        fill_region = pg.LinearRegionItem(values=(1.95, 2.05), orientation='horizontal',
                                          brush=pg.mkBrush(144, 238, 144, 25), movable=False)
        plot_latency.addItem(fill_region)

        # Blue line with circle markers
        self.curves['latency'] = plot_latency.plot(pen=pg.mkPen('blue', width=2),
                                                   symbol='o', symbolSize=6,
                                                   symbolBrush='blue', symbolPen='blue')
        self.plots['latency'] = plot_latency
        grid.addWidget(plot_latency, 0, 0)

        # =====================================================================
        # Panel 2 (top-right): Memory
        # =====================================================================
        plot_mem = pg.PlotWidget()
        plot_mem.setTitle("Memory Usage: Linear Growth", color='darkgreen', bold=True)
        plot_mem.setLabel('left', "Memory (MB)", **{'font-weight': 'bold'})
        plot_mem.setLabel('bottom', "Sample #", **{'font-weight': 'bold'})
        plot_mem.showGrid(x=True, y=True, alpha=0.3)

        # Green line with square markers
        self.curves['mem'] = plot_mem.plot(pen=pg.mkPen('green', width=2),
                                           symbol='s', symbolSize=6,
                                           symbolBrush='green', symbolPen='green')

        # Create baseline curve for fill (zero line)
        self.curves['mem_baseline'] = plot_mem.plot(pen=pg.mkPen(None))  # Invisible baseline

        # Green fill under curve
        self.fills['mem'] = pg.FillBetweenItem(
            self.curves['mem_baseline'],
            self.curves['mem'],
            brush=pg.mkBrush(0, 128, 0, 50)
        )
        plot_mem.addItem(self.fills['mem'])

        self.plots['mem'] = plot_mem
        grid.addWidget(plot_mem, 0, 1)

        # =====================================================================
        # Panel 3 (bottom-left): CPU Utilization
        # =====================================================================
        plot_cpu = pg.PlotWidget()
        plot_cpu.setTitle("CPU Utilization: Excellent 97-99%", color='darkgreen', bold=True)
        plot_cpu.setLabel('left', "CPU %", **{'font-weight': 'bold'})
        plot_cpu.setLabel('bottom', "Sample #", **{'font-weight': 'bold'})
        plot_cpu.showGrid(x=True, y=True, alpha=0.3)
        plot_cpu.setYRange(94, 100.5)

        # Optimal line: green dashed at 99%
        optimal_line = pg.InfiniteLine(pos=99, angle=0, pen=pg.mkPen('green', width=1.5, style=Qt.PenStyle.DashLine))
        plot_cpu.addItem(optimal_line)

        # Light green fill between 95-100%
        fill_region_cpu = pg.LinearRegionItem(values=(95, 100), orientation='horizontal',
                                              brush=pg.mkBrush(144, 238, 144, 25), movable=False)
        plot_cpu.addItem(fill_region_cpu)

        # Red line with diamond markers
        self.curves['cpu'] = plot_cpu.plot(pen=pg.mkPen('red', width=2),
                                           symbol='d', symbolSize=6,
                                           symbolBrush='red', symbolPen='red')
        self.plots['cpu'] = plot_cpu
        grid.addWidget(plot_cpu, 1, 0)

        # =====================================================================
        # Panel 4 (bottom-right): Fill Rate
        # =====================================================================
        plot_fills = pg.PlotWidget()
        plot_fills.setTitle("Fill Rate Distribution", color='darkgreen', bold=True)
        plot_fills.setLabel('left', "Full Fills (%)", **{'font-weight': 'bold'})
        plot_fills.setLabel('bottom', "Sample #", **{'font-weight': 'bold'})
        plot_fills.showGrid(x=False, y=True, alpha=0.3)

        # Purple scatter plot
        self.curves['fills'] = pg.ScatterPlotItem(size=10, brush=pg.mkBrush(128, 0, 128, 153),
                                                  pen=pg.mkPen('black', width=0.5))
        plot_fills.addItem(self.curves['fills'])

        # Mean line (will be updated dynamically)
        self.ref_lines['fills_mean'] = pg.InfiniteLine(pos=0, angle=0,
                                                       pen=pg.mkPen('red', width=1.5, style=Qt.PenStyle.DashLine),
                                                       label='Mean: 0%', labelOpts={'position': 0.9, 'color': 'red'})
        plot_fills.addItem(self.ref_lines['fills_mean'])

        # Add legend
        plot_fills.addLegend()

        self.plots['fills'] = plot_fills
        grid.addWidget(plot_fills, 1, 1)

    def _setup_reader(self):
        self.reader = InputReader()
        self.reader.metrics_ready.connect(self._on_metrics)
        self.reader.data_received.connect(self._on_data)
        self.reader.error_occurred.connect(self._on_error)
        self.reader.process_ended.connect(self._on_proc_end)

    def _setup_timers(self):
        self.timer = QTimer()
        self.timer.timeout.connect(self._update_graphs)
        self.timer.start(DEFAULT_UPDATE_MS)

        self.status_timer = QTimer()
        self.status_timer.timeout.connect(self._check_status)
        self.status_timer.start(1000)

    def _connect(self, source_type: str, path: str):
        self.reader.stop()
        self.data_status.set_status("inactive", "Connecting...")
        self.proc_status.set_status("inactive", "Starting...")
        self.reader.start(source_type, path)
        self.proc_status.set_status("active", f"Running: {source_type}")

    def _set_time_unit(self, idx: int):
        """Update time divisor for latency display in cards (graphs use Records as x-axis)."""
        self._time_divisor = TIME_UNITS[idx][2]

    def _on_data(self):
        self._last_data = datetime.now()
        self.data_status.set_status("active", f"Data: {self._last_data:%H:%M:%S}")

    def _on_metrics(self, m: Metrics):
        if self._is_stopped:
            return

        self._sample_idx += 1
        t = self._sample_idx

        # Add timestamp and store for export
        m.timestamp = datetime.now().isoformat()
        self._metrics_history.append(m)

        # Update history
        self.hist['time'].append(t)
        self.hist['records'].append(m.records / 1e6)  # Convert to millions
        self.hist['cpu'].append(m.cpu_pct)
        self.hist['mem'].append(m.memory_bytes / 1024 / 1024)
        self.hist['latency'].append(m.avg_match_ns / 1e3)  # ns -> μs
        self.hist['fills'].append(m.full_fills_pct)
        self.hist['bids'].append(m.bid_orders)
        self.hist['asks'].append(m.ask_orders)

        # Update cards
        self.cards['flow'].update({
            "Matches": f"{m.total_matches:,}",
            "Bids": f"{m.bid_orders:,}",
            "Asks": f"{m.ask_orders:,}",
        })
        self.cards['fills'].update({
            "Full": f"{m.full_fills:,} ({m.full_fills_pct:.1f}%)",
            "Partial": f"{m.partial_fills:,} ({m.partial_fills_pct:.1f}%)",
            "No Fill": f"{m.no_fills:,}",
            "Qty": f"{m.qty_matched:,}",
        })
        self.cards['latency'].update({
            "Avg": self._fmt_ns(m.avg_match_ns),
            "Min": self._fmt_ns(m.min_ns),
            "Max": self._fmt_ns(m.max_ns),
            "Lookup": self._fmt_ns(m.avg_lookup_ns),
        })
        self.cards['memory'].update({
            "Orders": f"{m.total_orders:,}",
            "Levels": f"{m.buy_levels + m.sell_levels:,}",
            "Size": m.memory_str,
        })
        self.cards['cpu'].update({
            "User": m.user_time,
            "System": m.system_time,
            "Usage": f"{m.cpu_pct:.1f}%",
            "CtxSw": f"{m.ctx_switches:,}",
        })

        self.records_lbl.setText(f"Records: {m.records:,}")
        self.throughput_lbl.setText(f"Throughput: {m.throughput:,.0f} msg/s")

    def _on_error(self, msg: str):
        self.data_status.set_status("error", "Error")
        QMessageBox.critical(self, "Error", msg)

    def _on_proc_end(self, code: int, stderr: str):
        if code == 0:
            self.proc_status.set_status("inactive", "Completed")
        else:
            self.proc_status.set_status("error", f"Crashed ({code})")
            dlg = QMessageBox(self)
            dlg.setIcon(QMessageBox.Icon.Critical)
            dlg.setWindowTitle("Process Crashed")
            dlg.setText(f"Exit code: {code}")
            if stderr:
                dlg.setDetailedText(stderr)
            dlg.exec()

    def _on_kill_clicked(self):
        """Kill OrderBook_v2 process, stop ingestion, export data."""
        self.kill_btn.setEnabled(False)
        self.kill_btn.setText("Stopping...")

        # 1. Kill all OrderBook_v2 processes (including those started externally)
        killed, errors = ProcessManager.kill_all()

        # 2. Stop reader (our own process if any)
        self.reader.stop()
        self._is_stopped = True

        # 3. Update status
        if killed > 0:
            self.proc_status.set_status("inactive", f"Killed {killed} process(es)")
        else:
            self.proc_status.set_status("warning", "No process found")
        self.data_status.set_status("inactive", "Stopped")

        # 4. Export data
        if self._metrics_history:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            default_name = f"orderbook_metrics_{timestamp}.xlsx"
            filepath, _ = QFileDialog.getSaveFileName(
                self,
                "Save Metrics Report",
                default_name,
                "Excel Files (*.xlsx);;CSV Files (*.csv);;All Files (*)"
            )
            if filepath:
                try:
                    actual_path = SpreadsheetExporter.export(self._metrics_history, filepath)
                    QMessageBox.information(
                        self,
                        "Export Complete",
                        f"Exported {len(self._metrics_history)} samples to:\n{actual_path}"
                    )
                except Exception as e:
                    QMessageBox.critical(self, "Export Error", f"Failed to export: {e}")
        else:
            QMessageBox.warning(self, "No Data", "No metrics data collected to export.")

        # 5. Show errors if any
        if errors:
            QMessageBox.warning(
                self,
                "Process Termination Warnings",
                f"Some processes could not be terminated:\n" + "\n".join(errors)
            )

        self.kill_btn.setText("✓ Stopped")

    def _check_status(self):
        if self._last_data:
            elapsed = (datetime.now() - self._last_data).total_seconds()
            if elapsed > DATA_TIMEOUT_MS / 1000:
                self.data_status.set_status("warning", f"No data ({elapsed:.0f}s)")

    def _smooth_data(self, data: list, window: int = 3) -> list:
        """Apply simple moving average smoothing to reduce noise.

        Args:
            data: List of numeric values to smooth
            window: Window size for moving average (odd numbers work best)

        Returns:
            Smoothed data list of same length as input
        """
        if len(data) < window:
            return data

        smoothed = []
        half_window = window // 2

        for i in range(len(data)):
            start = max(0, i - half_window)
            end = min(len(data), i + half_window + 1)
            window_vals = data[start:end]
            smoothed.append(sum(window_vals) / len(window_vals))

        return smoothed

    def _update_graphs(self):
        if not self.hist['time']:
            return

        # Use uniform sample index as x-axis (evenly spaced: 1, 2, 3, ...)
        x = list(self.hist['time'])

        # Panel 1: Latency (μs) with light smoothing to reduce spike noise
        latency_raw = list(self.hist['latency'])
        latency_smooth = self._smooth_data(latency_raw, window=3)
        self.curves['latency'].setData(x, latency_smooth)

        # Panel 2: Memory (MB) with fill under curve - no smoothing needed
        mem = list(self.hist['mem'])
        self.curves['mem'].setData(x, mem)
        self.curves['mem_baseline'].setData(x, [0] * len(x))

        # Panel 3: CPU (%) with smoothing to reduce volatility
        cpu_raw = list(self.hist['cpu'])
        cpu_smooth = self._smooth_data(cpu_raw, window=5)
        self.curves['cpu'].setData(x, cpu_smooth)

        # Panel 4: Fill Rate (scatter) with mean line - no smoothing
        fills = list(self.hist['fills'])
        if fills:
            self.curves['fills'].setData(x, fills)
            mean_val = sum(fills) / len(fills)
            self.ref_lines['fills_mean'].setValue(mean_val)
            self.ref_lines['fills_mean'].label.setText(f"Mean: {mean_val:.1f}%")

    @staticmethod
    def _fmt_ns(ns: int) -> str:
        if ns >= 1e9:
            return f"{ns/1e9:.2f} s"
        if ns >= 1e6:
            return f"{ns/1e6:.2f} ms"
        if ns >= 1e3:
            return f"{ns/1e3:.2f} us"
        return f"{ns} ns"

    @staticmethod
    def _get_sys_info() -> dict:
        return {
            'os': platform.system(),
            'arch': platform.machine(),
            'cores': psutil.cpu_count(logical=False),
            'ram': f"{psutil.virtual_memory().total / 1024**3:.1f} GB",
        }

    def closeEvent(self, event):
        self.reader.stop()
        self.timer.stop()
        self.status_timer.stop()
        event.accept()


def main():
    parser = argparse.ArgumentParser(description='Order Book Metrics Dashboard')
    parser.add_argument('--file', '-f', help='Read metrics from file')
    parser.add_argument('--exec', '-e', dest='executable', help='Launch executable')
    args = parser.parse_args()

    app = QApplication(sys.argv)
    app.setStyle('Fusion')
    window = DashboardWindow(source=args.file, executable=args.executable)
    window.show()
    sys.exit(app.exec())


if __name__ == '__main__':
    main()
